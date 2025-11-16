from django.shortcuts import render, redirect
from .forms import RegisterForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

# Create your views here.
def index(request):
    return render(request, 'index.html')
def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            return render(request, 'login.html', {'success': "Registration successful. Please login."})
        else:
            error_message = form.errors.as_text()
            return render(request, 'register.html', {'error': error_message})

    return render(request, 'register.html')

def login_view(request):
    if request.method=="POST":
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)
        if user is not None:
            login(request, user)
            return redirect("/dashboard")
        else:
            return render(request, 'login.html', {'error': "Invalid credentials. Please try again."})

    return render(request, 'login.html')

@login_required
def dashboard(request):
    return render(request, 'dashboard.html', {'name': request.user.first_name})

@login_required
def videocall(request):
    return render(request, 'videocall.html', {'name': request.user.first_name + " " + request.user.last_name})




@login_required
def logout_view(request):
    logout(request)
    return redirect("/login")

@login_required
def join_room(request):
    if request.method == 'POST':
        roomID = request.POST['roomID']
        return redirect("/meeting?roomID=" + roomID)
    return render(request, 'joinroom.html')

#for attandence 
# views.py

# from django.http import JsonResponse
# from openpyxl import Workbook, load_workbook
# from openpyxl.utils import get_column_letter
# import os
# from datetime import datetime

# def record_attendance(request):
#     if request.method == 'POST':
#         name = request.POST.get('name')

#         if not name:
#             return JsonResponse({'status': 'error', 'message': 'Name not provided'})

#         file_path = 'attendance.xlsx'
#         timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

#         if os.path.exists(file_path):
#             wb = load_workbook(file_path)
#             ws = wb.active
#         else:
#             wb = Workbook()
#             ws = wb.active
#             ws.append(['Name', 'Join Time'])  # Header row

#         ws.append([name, timestamp])
#         wb.save(file_path)

#         return JsonResponse({'status': 'success', 'message': 'Attendance recorded'})
#     else:
#         return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

from django.http import JsonResponse
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

def record_attendance(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        room_id = request.POST.get('roomID')  # ← Get roomID from frontend

        if not name or not room_id:
            return JsonResponse({'status': 'error', 'message': 'Name or Room ID not provided'})

        today = datetime.now().strftime('%Y-%m-%d')

        file_name = f"attendance_{room_id}_{today}.xlsx"
        file_path = os.path.join("attendance_files", file_name)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Name', 'Join Time'])

        ws.append([name, timestamp])
        wb.save(file_path)

        return JsonResponse({'status': 'success', 'message': f'Attendance recorded in {file_name}'})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



#--------------------------------------------------------------->>>>>>>>>>>>>>>>>>>

#for air canvas system
# from django.http import JsonResponse
# import threading
# from .aircanvas_runner import start_air_canvas

# def trigger_aircanvas(request):
#     threading.Thread(target=start_air_canvas).start()
#     return JsonResponse({'status': 'Air Canvas started'})

import subprocess
from django.http import JsonResponse
import os

# #start air canvas system
# def trigger_aircanvas(request):
#     script_path = os.path.join(os.path.dirname(__file__), 'aircanvas_runner.py')

#     # Launch in separate terminal (so GUI opens properly)
#     subprocess.Popen(['python', script_path])

#     return JsonResponse({'status': 'Air Canvas started'})




# canvas_process_pid = None
# #stoping air canvas system
# def stop_aircanvas(request):
#     global canvas_process_pid
#     if canvas_process_pid:
#         try:
#             os.kill(canvas_process_pid, signal.SIGTERM)
#             canvas_process_pid = None
#             return JsonResponse({'status': 'Air Canvas stopped'})
#         except Exception as e:
#             return JsonResponse({'status': 'Error stopping', 'error': str(e)})
#     else:
#         return JsonResponse({'status': 'No active Air Canvas'})


# ✅ Global process object
canvas_process = None

def trigger_aircanvas(request):
    global canvas_process
    if canvas_process is None or canvas_process.poll() is not None:
        canvas_process = subprocess.Popen(["python", "myapp/aircanvas_runner.py"])
        return JsonResponse({'status': 'Air Canvas started'})
    else:
        return JsonResponse({'status': 'Already running'})

def stop_aircanvas(request):
    global canvas_process
    if canvas_process and canvas_process.poll() is None:
        canvas_process.terminate()  # Sends SIGTERM
        canvas_process = None
        return JsonResponse({'status': 'Air Canvas stopped'})
    else:
        return JsonResponse({'status': 'No running canvas'})